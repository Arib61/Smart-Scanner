import os
import cv2
import pandas as pd
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Initialiser PaddleOCR en français
ocr = PaddleOCR(use_angle_cls=True, lang='fr')

def group_by_y(entries, tolerance=15):
    entries = sorted(entries, key=lambda e: e['y'])
    groups = []
    current = []
    last_y = None
    for entry in entries:
        if last_y is None or abs(entry['y'] - last_y) < tolerance:
            current.append(entry)
        else:
            groups.append(current)
            current = [entry]
        last_y = entry['y']
    if current:
        groups.append(current)
    return groups

def extract_table(image_path: str, min_confidence=0.5):
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image introuvable : {image_path}")

    result = ocr.ocr(image_path)

    entries = []

    for line in result[0]:
        try:
            box = line[0]

            # Cas normal : line[1] est un tuple (text, confidence)
            if isinstance(line[1], (list, tuple)) and len(line[1]) == 2:
                text, confidence = line[1]

            # Cas dégradé : line[1] est une liste (mais pas de confiance séparée)
            elif isinstance(line[1], (list, tuple)) and len(line[1]) == 1:
                text = line[1][0] if line[1][0] else ""
                confidence = 1.0

            # Cas extrême : juste un string ou vide
            elif isinstance(line[1], str):
                text = line[1]
                confidence = 1.0

            else:
                # Cas inconnu : ignorer
                continue

            if not text.strip() or confidence < min_confidence:
                continue

            x = min(pt[0] for pt in box)
            y = min(pt[1] for pt in box)
            entries.append({'text': text.strip(), 'x': x, 'y': y})

        except Exception as e:
            print(f"⚠️ Ligne ignorée à cause d'une erreur : {e}")
            continue




    lines = group_by_y(entries)
    table = []
    for line in lines:
        sorted_line = sorted(line, key=lambda e: e['x'])
        row = [cell['text'] for cell in sorted_line]
        table.append(row)

    max_cols = max(len(row) for row in table)
    table = [row + [""] * (max_cols - len(row)) for row in table]
    return table

def save_table_to_excel(table, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tableau OCR"

    # Styles
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for i, row in enumerate(table):
        for j, val in enumerate(row):
            cell = ws.cell(row=i+1, column=j+1, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill

    # Ajustement des colonnes
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path

def image_to_excel_converter_local(image_path: str, output_path: str,
                                   min_confidence: float = 0.5) -> str:
    print(f"🔍 Lecture de l'image : {image_path}")
    table = extract_table(image_path, min_confidence)
    if not table or all(len(row) <= 1 for row in table):
        raise ValueError("Aucun tableau structuré détecté dans l'image.")
    
    final_path = save_table_to_excel(table, output_path)
    print(f"✅ Export Excel terminé : {final_path}")
    return final_path
